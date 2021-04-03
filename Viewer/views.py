from django.shortcuts import render
from PropertyDocs.models import *
from Technical.models import *
from ImageUpload.models import *
from .BankExcelTemplates import *
from django.http import HttpResponse, HttpResponseRedirect, FileResponse
from django.core.exceptions import ObjectDoesNotExist
from django.db import models

db_models = [Documents,Address,Insights,MarketingValue,Plan,LegalLandmarks,
            SiteVisitLandmarks,Photos,Maps,AsPerDocuments,AsPerPlan,
            Actuals,Deviations,PermissibleBUA,ActualBUA,SanctionedArea,
            AreaDetails,Rate,TotalValue,FairMarketValue,PropertyStatus,FinalNotes
            ]



class ViewerManager(models.Manager):

    def document(self,DocID = None):

        dataSet = []
        pos = 0
        lookup = True

        while pos < len(db_models) and lookup:
            try:
                object = db_models[pos].objects.get(connection_id = DocID)
                dataSet.append(object)
                pos +=1

            except ObjectDoesNotExist:
                lookup = False
                print("LookUp Failed")
                dataSet = []

        if dataSet != []:

            bank = BankRef.objects.get(pk = DocID)
            customer = ClientInfo.objects.get(pk = bank.client_info_id)
            dataSet.insert(0,bank)
            dataSet.insert(0,customer)
            return dataSet


class ViewInterface(BankRef):
    viewer = ViewerManager()

def doc_complete(bank_id):
    """
    Document completion status is in relation Data Entry team
    Note: Data entry team doesn't enter technical detail data
    """
    try:
        #Check in final model
        if FinalNotes.objects.get(connection_id=bank_id):
            return True
    except FinalNotes.DoesNotExist:
        return False



#PDF Viewer controller
def ViewDocument(request,*args,**kwargs):

    if doc_complete(kwargs['bank_id']):
        """
        import openpyxl
        wb = openpyxl.Workbook('FileOperations/test.xlsx')
        wb.save('FileOperations/test.xlsx')
        load_workbook = openpyxl.load_workbook('FileOperations/test.xlsx')
        sheet = load_workbook.active
        sheet.title = "vinay"
        sheet['A1'] = 'Welcome to the future, Vini!!'
        load_workbook.save('FileOperations/test.xlsx')
        """

        input_file = 'FileOperations/test1.xlsx'

        import requests
        import base64
        import json
        
        excel_base64_data = None
        with open('FileOperations/test1.xlsx','rb') as excel:
            read_excel_base64 = excel.read()
            encode_excel = base64.b64encode(read_excel_base64)
            excel_base64_data = encode_excel.decode('utf-8')

        url = 'https://getoutpdf.com/api/convert/document-to-pdf'
        data = {
                "api_key": "32ffa3a2686cf2a8bc16d98541e9c2f0997de23636adc0d9df7b1b00b9da24b8",
                "document": excel_base64_data,
                }
        headers = {'Content-type': 'application/json', 'Accept': 'text/plain'}
        call_api_convertPdf = requests.post(url, data=json.dumps(data), headers=headers)
        print(call_api_convertPdf.json())

        pdf_base64_data = call_api_convertPdf.json()['pdf_base64']
        base64_img_bytes = pdf_base64_data.encode('utf-8')

        with open('FileOperations/decoded_image.pdf', 'wb') as pdf:
            decoded_image_data = base64.decodebytes(base64_img_bytes)
            pdf.write(decoded_image_data)

        with open('FileOperations/decoded_image.pdf', 'rb') as pdf:
            response = HttpResponse(pdf.read(),content_type='application/pdf')
            response['Content-Disposition'] = 'inline;filename=some_file.pdf'
            return response


    else:
        return HttpResponse("Document is In-complete")
