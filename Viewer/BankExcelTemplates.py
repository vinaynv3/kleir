import openpyxl
import pandas as pd
"""
Data representation & Manipulation classes
"""

class BFL_Urban:

    def __init__(self,data = None):
        self.data = data
        self.df = None

    def clean(self):

        data = {}
        for obj in self.data:
            data[str(obj.__class__).split(".")[-1].split("'")[0]] = []

            for k,v in obj.__dict__.items():
                if (not k.startswith("_")) and k != "id" and k != "connection_id":
                        data[str(obj.__class__).split(".")[-1].split("'")[0]].append(v)
        #Structure data into tabular form
        df = pd.DataFrame(dict([ (k,pd.Series(v)) for k,v in data.items() ]))
        return df


    """
    Excel cell points data structuring
    """
    def cleanData(self):
        self.df = self.clean()

    def personalDetailsContainer(self):

        df = self.df
        value_cell_points = ['A3','I4','C4','I5','C6','I6','D7','C8']

        # data is relative to value_cell_points position
        refNo = 'Reference Number : ' + df['BankRef'][1]
        data_tuple = (refNo,str(df['Documents'][1]),
                    df['Documents'][2],df['Documents'][3],
                    df['Documents'][4],df['ClientInfo'][3],
                    df['Documents'][5],df['Documents'][6])

        load_xlsx = openpyxl.load_workbook('BFL.xlsx')
        sheet = load_xlsx.active

        for i in range(len(value_cell_points)):
            sheet[value_cell_points[i]] = data_tuple[i]

        load_xlsx.save('BFL.xlsx')
        load_xlsx.close()
        del df

    def UpdateLocation(self):

        df = self.df
        value_cell_points = ['B10','B12','C14','B15','J15','B16','I16','C18','I18','D19']

        # data is relative to value_cell_points position
        data_tuple = (df['Address'][0],df['Address'][1],df['Address'][2],
                        df['Address'][3],df['Address'][4],df['MarketingValue'][0],
                        df['MarketingValue'][1],df['Insights'][0],df['MarketingValue'][2],
                        df['Insights'][1],)

        load_xlsx = openpyxl.load_workbook('BFL.xlsx')
        sheet = load_xlsx.active

        for i in range(len(value_cell_points)):
            sheet[value_cell_points[i]] = data_tuple[i]

        load_xlsx.save('BFL.xlsx')
        load_xlsx.close()
        del df

    def UpdatePropertyDesign(self):

        df = self.df
        value_cell_points = ['C25','J25','B31','I31','E33','C34']

        # data is relative to value_cell_points position
        data_tuple = (df['MarketingValue'][3],df['MarketingValue'][4],
                     df['Insights'][2],df['Insights'][3],df['Insights'][4],df['Insights'][5])

        load_xlsx = openpyxl.load_workbook('BFL.xlsx')
        sheet = load_xlsx.active

        for i in range(len(value_cell_points)):
            sheet[value_cell_points[i]] = data_tuple[i]

        load_xlsx.save('BFL.xlsx')
        load_xlsx.close()
        del df


    def VerifyDocs(self):

        df = self.df
        value_cell_points = ['B21','G21','B22','G22','B23','G23','B24','G24']

        # data is relative to value_cell_points position
        data_tuple = (df['LegalLandmarks'][0],df['SiteVisitLandmarks'][0],
                        df['LegalLandmarks'][1],df['SiteVisitLandmarks'][1],
                        df['LegalLandmarks'][2],df['SiteVisitLandmarks'][2],
                        df['LegalLandmarks'][3],df['SiteVisitLandmarks'][3])

        load_xlsx = openpyxl.load_workbook('BFL.xlsx')
        sheet = load_xlsx.active

        for i in range(len(value_cell_points)):
            sheet[value_cell_points[i]] = data_tuple[i]

        load_xlsx.save('BFL.xlsx')
        load_xlsx.close()
        del df

    def UpdatePropertyPlan(self):

        df = self.df
        value_cell_points = ['E26','E27','E28','D29','J29']

        # data is relative to value_cell_points position
        data_tuple = (df['Plan'][0],df['Plan'][1],df['Plan'][2],df['Plan'][3],df['Plan'][4],)

        load_xlsx = openpyxl.load_workbook('BFL.xlsx')
        sheet = load_xlsx.active

        for i in range(len(value_cell_points)):
            sheet[value_cell_points[i]] = data_tuple[i]

        load_xlsx.save('BFL.xlsx')
        load_xlsx.close()
        del df
